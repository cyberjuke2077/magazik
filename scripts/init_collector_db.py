"""Initialize collector database with products from Excel and CSV files."""

import sys
import csv
from pathlib import Path
from openpyxl import load_workbook
from collector.database import ProgressDB


def load_csv_products(csv_file: Path) -> list:
    """Load products from CSV file."""
    products = []
    
    # Try different encodings
    for encoding in ['gbk', 'utf-8', 'latin-1']:
        try:
            with open(csv_file, 'r', encoding=encoding) as f:
                reader = csv.reader(f)
                next(reader)  # Skip header
                
                for row in reader:
                    if row and row[0]:
                        part_number = str(row[0]).strip()
                        manufacturer = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        package = str(row[2]).strip() if len(row) > 2 and row[2] else None
                        
                        products.append({
                            'part_number': part_number,
                            'manufacturer': manufacturer,
                            'package': package
                        })
            
            return products
        except (UnicodeDecodeError, Exception):
            continue
    
    return products


def load_excel_products(excel_dir: str) -> list:
    """Load products from all Excel and CSV files."""
    products = []
    excel_path = Path(excel_dir)
    
    # Load Excel files
    for excel_file in sorted(excel_path.glob("*.xlsx")):
        print(f"Loading {excel_file.name}...")
        try:
            # Try data_only first
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active
            
            # Get max row (works even for corrupted files)
            max_row = ws.max_row
            
            # Skip header row
            for row_num in range(2, max_row + 1):
                row = [ws.cell(row_num, col).value for col in range(1, 4)]
                
                if row and row[0]:  # Part number exists
                    part_number = str(row[0]).strip()
                    manufacturer = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    package = str(row[2]).strip() if len(row) > 2 and row[2] else None
                    
                    products.append({
                        'part_number': part_number,
                        'manufacturer': manufacturer,
                        'package': package
                    })
            
            wb.close()
            print(f"  Loaded {excel_file.name}: {len(products)} products total")
        except Exception as e:
            print(f"  ⚠️  Error loading {excel_file.name}: {e}")
            continue
    
    # Load CSV files
    for csv_file in sorted(excel_path.glob("*.csv")):
        print(f"Loading {csv_file.name}...")
        try:
            csv_products = load_csv_products(csv_file)
            products.extend(csv_products)
            print(f"  Loaded {csv_file.name}: {len(products)} products total")
        except Exception as e:
            print(f"  ⚠️  Error loading {csv_file.name}: {e}")
            continue
    
    return products


def main():
    """Initialize database with products."""
    if len(sys.argv) < 2:
        print("Usage: python scripts/init_collector_db.py <excel_directory>")
        sys.exit(1)
    
    excel_dir = sys.argv[1]
    db_path = "./data/progress.db"
    
    print(f"Loading products from {excel_dir}...")
    products = load_excel_products(excel_dir)
    print(f"Found {len(products)} products")
    
    print(f"Initializing database at {db_path}...")
    db = ProgressDB(db_path)
    
    print("Adding products to queue...")
    for i, product in enumerate(products, 1):
        db.add_product(
            product['part_number'],
            product['manufacturer'],
            product['package']
        )
        
        if i % 1000 == 0:
            print(f"  Added {i}/{len(products)} products...")
    
    stats = db.get_stats()
    print(f"\nDatabase initialized successfully!")
    print(f"Total products: {stats['total']}")
    print(f"Pending: {stats['pending']}")
    
    db.close()


if __name__ == "__main__":
    main()
